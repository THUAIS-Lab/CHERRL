"""Re-export manual_audit_sheet_{A,B}.{csv,xlsx} from the hidden CSV.

Fixes prior export issues by:
  - encoding utf-8-sig
  - csv.QUOTE_ALL
  - lineterminator='\\n'
  - replacing embedded \\r / \\n in prompt_input / model_output with the
    literal two-char string "\\n" so each sample occupies exactly one CSV line
  - also producing .xlsx with column widths and wrap-text

Self-checks every output and prints a summary. Does NOT touch the hidden CSV
or re-sample.

Usage:
  python export_sheets_strict.py
  python export_sheets_strict.py --out /custom/dir
"""

from __future__ import annotations

import argparse, csv, sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SHEET_COLS = [
    "sample_id",
    "run_name",
    "target_shortcut_definition",
    "prompt_input",
    "model_output",
    "label_0_1_2",
    "optional_note",
]

FORBIDDEN_COLS = {
    "step", "score", "actual_step", "visible_score", "region",
    "window_start", "window_end", "run_id", "source_file_or_row_id",
    "notes", "reference", "detector", "threshold", "canonical",
    "interval", "onset", "L", "U", "C",
}

TEXT_COLS_NEEDING_NEWLINE_FLATTEN = ("prompt_input", "model_output")


def flatten_newlines(s: str) -> str:
    if s is None: return ""
    s = str(s)
    # Replace CRLF first, then bare CR / LF with literal "\n" so each row
    # occupies exactly one CSV line. Don't touch other whitespace.
    return s.replace("\r\n", "\\n").replace("\n", "\\n").replace("\r", "\\n")


def build_sheet_df(hidden_df: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame()
    out["sample_id"] = hidden_df["sample_id"].astype(str)
    out["run_name"] = hidden_df["run_name"].astype(str)
    out["target_shortcut_definition"] = hidden_df["target_shortcut_definition"].astype(str)
    out["prompt_input"] = hidden_df["prompt_input"].astype(str).map(flatten_newlines)
    out["model_output"] = hidden_df["model_output"].astype(str).map(flatten_newlines)
    out["label_0_1_2"] = ""
    out["optional_note"] = ""
    # Confirm column order
    out = out[SHEET_COLS]
    return out


def write_csv(df: pd.DataFrame, path: Path):
    df.to_csv(
        path,
        index=False,
        encoding="utf-8-sig",
        quoting=csv.QUOTE_ALL,
        lineterminator="\n",
    )


def write_xlsx(df: pd.DataFrame, path: Path, sheet_name: str = "manual_audit"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Body
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val if val != "" else None)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths: tuned for readability
    widths = {
        "sample_id": 22,
        "run_name": 24,
        "target_shortcut_definition": 60,
        "prompt_input": 80,
        "model_output": 80,
        "label_0_1_2": 12,
        "optional_note": 36,
    }
    for j, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(j)].width = widths.get(col, 20)

    # Freeze header
    ws.freeze_panes = "A2"
    wb.save(path)


def self_check(path: Path, expected_n: int) -> dict:
    df = pd.read_csv(path, encoding="utf-8-sig")
    issues = []
    if df.shape != (expected_n, len(SHEET_COLS)):
        issues.append(f"shape mismatch: got {df.shape}, expected ({expected_n}, {len(SHEET_COLS)})")
    if list(df.columns) != SHEET_COLS:
        issues.append(f"columns mismatch: got {list(df.columns)}")
    if any(str(c).startswith("Unnamed") for c in df.columns):
        issues.append(f"contains Unnamed columns: {list(df.columns)}")
    leaks = [c for c in df.columns if c.lower() in FORBIDDEN_COLS]
    if leaks: issues.append(f"forbidden hidden fields leaked: {leaks}")
    n_unique = df["sample_id"].nunique()
    if n_unique != expected_n:
        issues.append(f"sample_id not unique: {n_unique}/{expected_n}")
    # label / note must be empty
    nonempty_label = (df["label_0_1_2"].fillna("").astype(str).str.strip() != "").sum()
    nonempty_note = (df["optional_note"].fillna("").astype(str).str.strip() != "").sum()
    if nonempty_label: issues.append(f"label_0_1_2 has {nonempty_label} prefilled values")
    if nonempty_note:  issues.append(f"optional_note has {nonempty_note} prefilled values")
    # per-run counts
    run_counts = df["run_name"].value_counts().to_dict()
    return {
        "shape": df.shape,
        "columns": list(df.columns),
        "unique_sample_ids": n_unique,
        "run_name_counts": run_counts,
        "label_0_1_2_nonempty": int(nonempty_label),
        "optional_note_nonempty": int(nonempty_note),
        "issues": issues,
    }


def main():
    here = Path(__file__).resolve().parent
    default_dir = here.parent

    ap = argparse.ArgumentParser()
    ap.add_argument("--out", type=Path, default=default_dir,
                    help="manual_audit/ directory containing hidden CSV")
    args = ap.parse_args()

    out_dir = args.out
    hidden_p = out_dir / "manual_audit_samples_hidden.csv"
    if not hidden_p.exists():
        print(f"ERROR: hidden CSV not found: {hidden_p}", file=sys.stderr); sys.exit(1)

    # Hidden file may contain embedded newlines in prompt_input/model_output;
    # those were properly quoted in the original write (csv.DictWriter default
    # uses QUOTE_MINIMAL). pandas handles multi-line quoted fields out of the box.
    hidden_df = pd.read_csv(hidden_p, encoding="utf-8", engine="python")
    print(f"hidden: rows={len(hidden_df)} cols={list(hidden_df.columns)}")
    assert len(hidden_df) == 180, f"hidden has {len(hidden_df)} rows, expected 180"

    # Verify per (run_id, region) count = 10
    per_rr = hidden_df.groupby(["run_id", "region"]).size().to_dict()
    bad = [k for k, v in per_rr.items() if v != 10]
    assert not bad, f"per (run_id, region) != 10: {bad}"
    print(f"per (run_id, region) all == 10  ✓")

    sheet_df = build_sheet_df(hidden_df)

    written = []
    for letter in ("A", "B"):
        csv_p = out_dir / f"manual_audit_sheet_{letter}.csv"
        xlsx_p = out_dir / f"manual_audit_sheet_{letter}.xlsx"
        write_csv(sheet_df, csv_p); written.append(csv_p)
        write_xlsx(sheet_df, xlsx_p, sheet_name=f"manual_audit_{letter}")
        written.append(xlsx_p)
        print(f"wrote {csv_p}")
        print(f"wrote {xlsx_p}")

    # Self-check both CSVs (XLSX correctness verified by openpyxl write API)
    reports = {}
    for letter in ("A", "B"):
        csv_p = out_dir / f"manual_audit_sheet_{letter}.csv"
        rep = self_check(csv_p, expected_n=180)
        reports[letter] = rep
        print(f"\n=== self-check sheet_{letter}.csv ===")
        print(f"  shape: {rep['shape']}")
        print(f"  columns: {rep['columns']}")
        print(f"  unique sample_ids: {rep['unique_sample_ids']}")
        print(f"  label_0_1_2 prefilled: {rep['label_0_1_2_nonempty']}")
        print(f"  optional_note prefilled: {rep['optional_note_nonempty']}")
        print(f"  run_name counts: {rep['run_name_counts']}")
        print(f"  issues: {rep['issues'] or 'NONE ✓'}")

    # Same set of sample_ids between A and B
    a_df = pd.read_csv(out_dir / "manual_audit_sheet_A.csv", encoding="utf-8-sig")
    b_df = pd.read_csv(out_dir / "manual_audit_sheet_B.csv", encoding="utf-8-sig")
    same_ids = set(a_df["sample_id"]) == set(b_df["sample_id"])
    print(f"\nA and B sample_id set match: {same_ids} ✓" if same_ids else "MISMATCH ✗")

    # XLSX read-back check
    for letter in ("A", "B"):
        xlsx_p = out_dir / f"manual_audit_sheet_{letter}.xlsx"
        x_df = pd.read_excel(xlsx_p, engine="openpyxl")
        ok = x_df.shape == (180, len(SHEET_COLS)) and list(x_df.columns) == SHEET_COLS
        print(f"xlsx_{letter}: shape={x_df.shape} columns_ok={list(x_df.columns)==SHEET_COLS}  {'✓' if ok else '✗'}")

    print(f"\nALL DONE. files: {[p.name for p in written]}")


if __name__ == "__main__":
    main()
